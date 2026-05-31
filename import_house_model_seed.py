from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
import re

POSSIBLE_PATHS = [
    Path("inputs/House Model Data.xlsx"),
    Path("House Model Data.xlsx"),
]

OUTPUT_PATH = Path("inputs/house_race_inputs.csv")

EXPECTED_COLUMNS = [
    "State",
    "District",
    "Incumbent",
    "2024 Margin",
    "2020 Margin",
    "GenBallot Adjusted Margin",
    "Dem Candidate",
    "GOP Candidate",
]

OPTIONAL_COLUMNS = [
    "Incumbent Party",
    "Region",
    "District Type",
    "State Environment Adjustment",
    "College Share Tier",
    "White Share Tier",
    "Black Share Tier",
    "Hispanic Share Tier",
    "Median Income Tier",
]

AT_LARGE_STATES = {"AK", "DE", "ND", "SD", "VT", "WY"}


def find_workbook_path():
    for path in POSSIBLE_PATHS:
        if path.exists():
            return path

    raise FileNotFoundError(
        "Could not find House Model Data.xlsx. Put it in either "
        "inputs/House Model Data.xlsx or the project root."
    )


def clean_text(x):
    if x is None:
        return ""
    return str(x).strip()


def parse_margin(value):
    if value is None or value == "":
        return None

    if isinstance(value, (int, float)):
        return float(value)

    s = str(value).strip().upper()

    if s in ["", "NAN", "NONE"]:
        return None

    s = s.replace("DEM", "D")
    s = s.replace("DEMOCRAT", "D")
    s = s.replace("DEMOCRATIC", "D")
    s = s.replace("REP", "R")
    s = s.replace("GOP", "R")
    s = s.replace("REPUBLICAN", "R")
    s = s.replace(" ", "")

    m = re.match(r"^D\+?(-?\d+(\.\d+)?)$", s)
    if m:
        return float(m.group(1))

    m = re.match(r"^R\+?(-?\d+(\.\d+)?)$", s)
    if m:
        return -float(m.group(1))

    try:
        return float(s)
    except Exception:
        return None


def normalize_incumbent_party(value):
    s = clean_text(value).upper()

    if s in ["D", "DEM", "DEMOCRAT", "DEMOCRATIC"]:
        return "D"

    if s in ["R", "REP", "REPUBLICAN", "GOP"]:
        return "R"

    if s in ["I", "IND", "INDEPENDENT"]:
        return "I"

    if s in ["VACANT", "OPEN", "NONE", "NO INCUMBENT"]:
        return "Vacant"

    if s == "":
        return ""

    return s

def normalize_region(value):
    s = clean_text(value)

    if s == "":
        return "Unknown Region"

    aliases = {
        "northeast": "Northeast",
        "mid-atlantic": "Mid-Atlantic",
        "mid atlantic": "Mid-Atlantic",
        "deep south": "Deep South",
        "middle south": "Middle South",
        "urban south": "Urban South",
        "appalachia": "Appalachia",
        "midwest": "Midwest",
        "great plains": "Great Plains",
        "mountain west": "Mountain West",
        "pacific": "Pacific",
        "northwest": "Northwest",
        "unknown region": "Unknown Region",
    }

    return aliases.get(s.lower(), s)


def normalize_district_type(value):
    s = clean_text(value)

    if s == "":
        return "Mixed"

    aliases = {
        "urban": "Urban",
        "suburban": "Suburban",
        "exurban": "Exurban",
        "rural": "Rural",
        "mixed": "Mixed",
    }

    return aliases.get(s.lower(), s)


def normalize_tier(value, default="Unknown"):
    s = clean_text(value)

    if s == "":
        return default

    aliases = {
        "low": "Low",
        "medium": "Medium",
        "med": "Medium",
        "high": "High",
        "very high": "Very High",
        "very_high": "Very High",
        "veryhigh": "Very High",
        "unknown": "Unknown",
    }

    return aliases.get(s.lower(), s)


def normalize_district_value(state, district):
    state = clean_text(state).upper()
    district = clean_text(district).upper()

    if district in ["", "NAN", "NONE"]:
        return ""

    if state in AT_LARGE_STATES and district in ["1", "01", "AL", "AT-LARGE", "AT LARGE", "AT_LARGE"]:
        return "AL"

    if district.isdigit():
        return str(int(district))

    return district


def normalize_district_id(state, district):
    state = clean_text(state).upper()
    district = normalize_district_value(state, district)

    if state == "" or district == "":
        return ""

    return f"{state}-{district}"


def optional_cell(ws, row, col_index, column_name):
    if column_name not in col_index:
        return None
    return ws.cell(row, col_index[column_name]).value


def cell_is_italic(cell):
    try:
        return bool(cell.font and cell.font.italic)
    except Exception:
        return False


def main():
    input_path = find_workbook_path()

    wb = load_workbook(input_path, data_only=True)
    ws = wb.active

    headers = [
        clean_text(ws.cell(row=1, column=c).value)
        for c in range(1, ws.max_column + 1)
    ]

    missing = [c for c in EXPECTED_COLUMNS if c not in headers]

    if missing:
        raise ValueError(
            "Workbook is missing expected columns: "
            + ", ".join(missing)
            + "\nFound columns: "
            + ", ".join(headers)
        )

    col_index = {name: headers.index(name) + 1 for name in headers}

    optional_found = [c for c in OPTIONAL_COLUMNS if c in headers]
    optional_missing = [c for c in OPTIONAL_COLUMNS if c not in headers]

    print("Optional columns found:", ", ".join(optional_found) if optional_found else "None")
    if optional_missing:
        print("Optional columns missing:", ", ".join(optional_missing))

    rows = []

    for r in range(2, ws.max_row + 1):
        state = clean_text(ws.cell(r, col_index["State"]).value).upper()
        district_raw = clean_text(ws.cell(r, col_index["District"]).value)
        district = normalize_district_value(state, district_raw)
        district_id = normalize_district_id(state, district)

        if state == "" and district == "":
            continue

        incumbent_raw = clean_text(ws.cell(r, col_index["Incumbent"]).value)

        incumbent_party_raw = clean_text(
            optional_cell(ws, r, col_index, "Incumbent Party")
        )

        dem_cell = ws.cell(r, col_index["Dem Candidate"])
        gop_cell = ws.cell(r, col_index["GOP Candidate"])

        dem_candidate = clean_text(dem_cell.value)
        gop_candidate = clean_text(gop_cell.value)

        dem_candidate_italic = cell_is_italic(dem_cell)
        gop_candidate_italic = cell_is_italic(gop_cell)

        # Workbook convention: non-incumbents are italicized.
        dem_candidate_is_incumbent = bool(dem_candidate and not dem_candidate_italic)
        gop_candidate_is_incumbent = bool(gop_candidate and not gop_candidate_italic)

        # Prefer the explicit Incumbent Party column when present.
        # Fall back to the older Incumbent field only if Incumbent Party is blank.
        incumbent_party = normalize_incumbent_party(incumbent_party_raw)

        if incumbent_party == "":
            incumbent_party = normalize_incumbent_party(incumbent_raw)

        if dem_candidate_is_incumbent:
            inferred_incumbent_party = "D"
        elif gop_candidate_is_incumbent:
            inferred_incumbent_party = "R"
        else:
            inferred_incumbent_party = incumbent_party

        incumbent_running = dem_candidate_is_incumbent or gop_candidate_is_incumbent
        open_seat = not incumbent_running

        pres_2024_margin_dem = parse_margin(ws.cell(r, col_index["2024 Margin"]).value)
        pres_2020_margin_dem = parse_margin(ws.cell(r, col_index["2020 Margin"]).value)
        genballot_adjusted_margin_dem = parse_margin(
            ws.cell(r, col_index["GenBallot Adjusted Margin"]).value
        )

        region = normalize_region(optional_cell(ws, r, col_index, "Region"))
        district_type = normalize_district_type(optional_cell(ws, r, col_index, "District Type"))

        state_environment_adjustment_dem = parse_margin(
            optional_cell(ws, r, col_index, "State Environment Adjustment")
        )
        if state_environment_adjustment_dem is None:
            state_environment_adjustment_dem = 0.0

        college_share_tier = normalize_tier(
            optional_cell(ws, r, col_index, "College Share Tier"),
            default="Unknown"
        )

        white_share_tier = normalize_tier(
            optional_cell(ws, r, col_index, "White Share Tier"),
            default="Unknown"
        )

        black_share_tier = normalize_tier(
            optional_cell(ws, r, col_index, "Black Share Tier"),
            default="Unknown"
        )

        hispanic_share_tier = normalize_tier(
            optional_cell(ws, r, col_index, "Hispanic Share Tier"),
            default="Unknown"
        )

        median_income_tier = normalize_tier(
            optional_cell(ws, r, col_index, "Median Income Tier"),
            default="Unknown"
        )

        education_race_error_group = f"{college_share_tier} College / {white_share_tier} White"

        demographic_error_group = (
            f"{college_share_tier} College / "
            f"{white_share_tier} White / "
            f"{black_share_tier} Black / "
            f"{hispanic_share_tier} Hispanic / "
            f"{median_income_tier} Income"
        )

        rows.append(
            {
                "state": state,
                "district": district,
                "district_id": district_id,

                "region": region,
                "district_type": district_type,
                "college_share_tier": college_share_tier,
                "white_share_tier": white_share_tier,
                "black_share_tier": black_share_tier,
                "hispanic_share_tier": hispanic_share_tier,
                "median_income_tier": median_income_tier,
                "education_race_error_group": education_race_error_group,
                "demographic_error_group": demographic_error_group,

                "state_error_group": state,
                "region_error_group": region,
                "district_type_error_group": district_type,
                "state_environment_adjustment_dem": state_environment_adjustment_dem,

                "incumbent_raw": incumbent_raw,
                "incumbent_party_raw": incumbent_party_raw,
                "incumbent_party": incumbent_party,
                "inferred_incumbent_party": inferred_incumbent_party,
                "incumbent_running": incumbent_running,
                "open_seat": open_seat,

                "dem_candidate": dem_candidate,
                "gop_candidate": gop_candidate,

                "dem_candidate_italic": dem_candidate_italic,
                "gop_candidate_italic": gop_candidate_italic,
                "dem_candidate_is_incumbent": dem_candidate_is_incumbent,
                "gop_candidate_is_incumbent": gop_candidate_is_incumbent,

                "pres_2024_margin_dem": pres_2024_margin_dem,
                "pres_2020_margin_dem": pres_2020_margin_dem,
                "genballot_adjusted_margin_dem": genballot_adjusted_margin_dem,

                "district_partisan_baseline_dem": None,
                "district_elasticity": 0.90,
                "manual_elasticity_override": None,
                "national_environment_margin_dem": None,
                "district_environment_adjustment_dem": None,

                "incumbency_adjustment_dem": None,
                "candidate_quality_adjustment_dem": 0.0,
                "special_adjustment_dem": 0.0,

                "fundamentals_margin_dem": None,
                "polling_margin_dem": None,
                "poll_count": 0,
                "polling_active": False,
                "model_margin_dem": None,
                "dem_win_probability": None,

                "race_notes": "",
            }
        )

    df = pd.DataFrame(rows)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(OUTPUT_PATH, index=False)

    print(f"Imported {len(df)} House races from {input_path}")
    print(f"Wrote {OUTPUT_PATH}")

    print()
    print("Incumbency summary:")
    print(df[
        [
            "incumbent_running",
            "open_seat",
            "dem_candidate_is_incumbent",
            "gop_candidate_is_incumbent",
        ]
    ].sum(numeric_only=True).to_string())

    print()
    print("Demographic / grouping summary:")
    for col in [
        "region",
        "district_type",
        "college_share_tier",
        "white_share_tier",
        "black_share_tier",
        "hispanic_share_tier",
        "median_income_tier",
        "education_race_error_group",
    ]:
        print()
        print(col)
        print(df[col].value_counts(dropna=False).head(20).to_string())

    print()
    print("First 20 rows:")
    preview_cols = [
        "district_id",
        "incumbent_raw",
        "incumbent_party_raw",
        "incumbent_party",
        "inferred_incumbent_party",
        "region",
        "district_type",
        "college_share_tier",
        "white_share_tier",
        "black_share_tier",
        "hispanic_share_tier",
        "median_income_tier",
        "education_race_error_group",
        "dem_candidate",
        "gop_candidate",
        "dem_candidate_is_incumbent",
        "gop_candidate_is_incumbent",
        "pres_2024_margin_dem",
        "pres_2020_margin_dem",
    ]
    print(df[preview_cols].head(20).to_string(index=False))

    if len(df) != 435:
        print()
        print(f"WARNING: expected 435 rows, imported {len(df)} rows.")


if __name__ == "__main__":
    main()
